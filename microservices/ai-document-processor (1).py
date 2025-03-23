import os
import sys
import logging
import json
import time
from typing import List, Dict, Any, Optional, Tuple, Union
from pathlib import Path
import traceback

import numpy as np
import pandas as pd
import torch
from transformers import (
    AutoTokenizer, 
    AutoModelForTokenClassification,
    AutoModelForSequenceClassification,
    AutoModelForQuestionAnswering,
    pipeline
)
from sentence_transformers import SentenceTransformer
import spacy
import pytesseract
from pdf2image import convert_from_path, convert_from_bytes
from pdfminer.high_level import extract_text
import docx2txt
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import fitz  # PyMuPDF

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class AIDocumentProcessor:
    """
    AI-based document processing module for the Advanced Pilot Training Platform.
    
    This module provides capabilities for document understanding, text extraction,
    entity recognition, and relationship extraction from training documents.
    """
    
    def __init__(self, model_dir: str = "./models", use_gpu: bool = True):
        """
        Initialize the AI document processor with the specified models.
        
        Args:
            model_dir: Directory containing pre-trained models
            use_gpu: Whether to use GPU for inference if available
        """
        self.model_dir = Path(model_dir)
        self.use_gpu = use_gpu and torch.cuda.is_available()
        self.device = torch.device("cuda" if self.use_gpu else "cpu")
        
        logger.info(f"Initializing AI Document Processor (using {'GPU' if self.use_gpu else 'CPU'})")
        
        # Load NLP components
        self._load_models()
    
    def _load_models(self):
        """Load all required NLP models."""
        try:
            # Load NER model for aviation/training domain entities
            ner_model_path = self.model_dir / "ner"
            if ner_model_path.exists():
                self.ner_tokenizer = AutoTokenizer.from_pretrained(ner_model_path)
                self.ner_model = AutoModelForTokenClassification.from_pretrained(ner_model_path).to(self.device)
                self.ner_pipeline = pipeline("ner", model=self.ner_model, tokenizer=self.ner_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic NER model
                logger.warning(f"Custom NER model not found at {ner_model_path}, using default model")
                self.ner_pipeline = pipeline("ner", model="dbmdz/bert-large-cased-finetuned-conll03-english", device=0 if self.use_gpu else -1)
            
            # Load document classification model
            classifier_model_path = self.model_dir / "classifier"
            if classifier_model_path.exists():
                self.classifier_tokenizer = AutoTokenizer.from_pretrained(classifier_model_path)
                self.classifier_model = AutoModelForSequenceClassification.from_pretrained(classifier_model_path).to(self.device)
                self.classifier_pipeline = pipeline("text-classification", model=self.classifier_model, tokenizer=self.classifier_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic classification model
                logger.warning(f"Custom classifier model not found at {classifier_model_path}, using default model")
                self.classifier_pipeline = pipeline("text-classification", model="distilbert-base-uncased-finetuned-sst-2-english", device=0 if self.use_gpu else -1)
            
            # Load QA model for relation extraction
            qa_model_path = self.model_dir / "qa"
            if qa_model_path.exists():
                self.qa_tokenizer = AutoTokenizer.from_pretrained(qa_model_path)
                self.qa_model = AutoModelForQuestionAnswering.from_pretrained(qa_model_path).to(self.device)
                self.qa_pipeline = pipeline("question-answering", model=self.qa_model, tokenizer=self.qa_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic QA model
                logger.warning(f"Custom QA model not found at {qa_model_path}, using default model")
                self.qa_pipeline = pipeline("question-answering", model="deepset/roberta-base-squad2", device=0 if self.use_gpu else -1)
            
            # Load sentence embedding model for similarity and clustering
            self.sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
            if self.use_gpu:
                self.sentence_model = self.sentence_model.to(self.device)
            
            # Load spaCy for text processing
            try:
                self.nlp = spacy.load("en_core_web_trf")
            except OSError:
                logger.warning("Could not load en_core_web_trf. Downloading now...")
                spacy.cli.download("en_core_web_trf")
                self.nlp = spacy.load("en_core_web_trf")
            
            logger.info("All NLP models loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading models: {str(e)}")
            logger.error(traceback.format_exc())
            raise RuntimeError(f"Failed to initialize AI Document Processor: {str(e)}")
    
    def process_document(self, file_path: Union[str, Path], 
                        options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Process a document file and extract structured information.
        
        Args:
            file_path: Path to the document file
            options: Processing options and parameters
        
        Returns:
            Dictionary containing processed document data
        """
        if options is None:
            options = {}
        
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Document file not found: {file_path}")
        
        try:
            # Determine document type and extract text
            doc_type, text = self.extract_text(file_path)
            
            # Process the text
            doc_info = {
                "document_type": doc_type,
                "text": text[:1000] + "..." if len(text) > 1000 else text,  # Truncate for result object
                "file_path": str(file_path),
                "file_name": file_path.name,
                "file_size": file_path.stat().st_size,
                "processing_time": time.time(),
            }
            
            # Document classification
            if options.get("classify_document", True):
                doc_info["document_classification"] = self.classify_document(text)
            
            # Extract entities
            if options.get("extract_entities", True):
                doc_info["entities"] = self.extract_entities(text)
            
            # Extract relationships
            if options.get("extract_relationships", True):
                doc_info["relationships"] = self.extract_relationships(text, doc_info.get("entities", []))
            
            # Extract regulatory references
            if options.get("extract_regulations", True):
                doc_info["regulatory_references"] = self.extract_regulatory_references(text)
            
            # Extract training elements
            if options.get("extract_training_elements", True):
                doc_info["training_elements"] = self.extract_training_elements(text)
            
            # Extract document structure
            if options.get("extract_structure", True):
                doc_info["structure"] = self.extract_document_structure(text)
            
            # Generate summary
            if options.get("generate_summary", True):
                doc_info["summary"] = self.generate_summary(text)
            
            return doc_info
            
        except Exception as e:
            logger.error(f"Error processing document {file_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise RuntimeError(f"Failed to process document: {str(e)}")
    
    def extract_text(self, file_path: Path) -> Tuple[str, str]:
        """
        Extract text from a document file based on its type.
        
        Args:
            file_path: Path to the document file
        
        Returns:
            Tuple of (document_type, extracted_text)
        """
        suffix = file_path.suffix.lower()
        
        if suffix == '.pdf':
            return "pdf", self._extract_from_pdf(file_path)
        elif suffix in ['.docx', '.doc']:
            return "docx", self._extract_from_docx(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return "excel", self._extract_from_excel(file_path)
        elif suffix in ['.html', '.htm']:
            return "html", self._extract_from_html(file_path)
        elif suffix == '.txt':
            return "text", file_path.read_text(errors='replace')
        elif suffix in ['.pptx', '.ppt']:
            return "powerpoint", self._extract_from_ppt(file_path)
        else:
            # Try to extract as text as a fallback
            try:
                return "unknown", file_path.read_text(errors='replace')
            except Exception:
                raise ValueError(f"Unsupported document type: {suffix}")
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from a PDF file using multiple methods for best results."""
        text = ""
        
        # Try pdfminer first (better for text-based PDFs)
        try:
            text = extract_text(str(file_path))
        except Exception as e:
            logger.warning(f"pdfminer extraction failed: {str(e)}")
        
        # If text is too short, try PyMuPDF
        if len(text.strip()) < 100:
            try:
                doc = fitz.open(str(file_path))
                pages_text = []
                for page in doc:
                    pages_text.append(page.get_text())
                text = "\n\n".join(pages_text)
                doc.close()
            except Exception as e:
                logger.warning(f"PyMuPDF extraction failed: {str(e)}")
        
        # If still too short, try OCR
        if len(text.strip()) < 100:
            try:
                images = convert_from_path(str(file_path))
                ocr_texts = []
                for img in images:
                    ocr_texts.append(pytesseract.image_to_string(img))
                text = "\n\n".join(ocr_texts)
            except Exception as e:
                logger.warning(f"OCR extraction failed: {str(e)}")
        
        return text
    
    def _extract_from_docx(self, file_path: Path) -> str:
        """Extract text from a Word document."""
        return docx2txt.process(str(file_path))
    
    def _extract_from_excel(self, file_path: Path) -> str:
        """Extract text from an Excel spreadsheet."""
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        texts = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            texts.append(f"Sheet: {sheet}")
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            
            texts.append("\n".join(rows))
        
        wb.close()
        return "\n\n".join(texts)
    
    def _extract_from_html(self, file_path: Path) -> str:
        """Extract text from an HTML file."""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            return soup.get_text(separator="\n")
    
    def _extract_from_ppt(self, file_path: Path) -> str:
        """Extract text from a PowerPoint presentation."""
        # This is a placeholder - in a real implementation we would use
        # python-pptx or a similar library
        return f"PowerPoint extraction not fully implemented for {file_path}"
    
    def classify_document(self, text: str) -> Dict[str, Any]:
        """
        Classify the document type and category.
        
        Args:
            text: Document text
        
        Returns:
            Dictionary with classification results
        """
        # Use only first 1000 tokens for classification
        truncated_text = " ".join(text.split()[:1000])
        
        # Get predicted categories
        result = self.classifier_pipeline(truncated_text)
        
        # Convert result to our expected format
        if isinstance(result, list):
            result = result[0]
        
        return {
            "category": result["label"],
            "confidence": float(result["score"])
        }
    
    def extract_entities(self, text: str) -> List[Dict[str, Any]]:
        """
        Extract named entities from document text.
        
        Args:
            text: Document text
        
        Returns:
            List of extracted entities with metadata
        """
        # Use spaCy for initial entity extraction
        doc = self.nlp(text[:10000])  # Limit size for performance
        
        entities = []
        for ent in doc.ents:
            entities.append({
                "text": ent.text,
                "label": ent.label_,
                "start": ent.start_char,
                "end": ent.end_char,
                "confidence": 0.9  # Placeholder, spaCy doesn't provide confidence scores
            })
        
        # Extract domain-specific entities using the specialized NER model
        # Process text in chunks of 512 tokens to avoid exceeding model length limits
        tokenized = self.ner_tokenizer(text, return_offsets_mapping=True, truncation=True)
        offset_mapping = tokenized["offset_mapping"]
        
        specialized_results = self.ner_pipeline(text[:10000])
        
        # Process and deduplicate NER results
        seen_spans = set()
        for item in specialized_results:
            # Map subtoken indices to character spans
            if 'start' in item and 'end' in item:
            # Map subtoken indices to character spans
            if 'start' in item and 'end' in item:
                entity_text = text[item['start']:item['end']]
                span_key = (item['start'], item['end'])
            else:
                # Handle word_index format
                start_idx = offset_mapping[item['word_index']][0]
                # Handle BPE tokenization where end might be in the next token
                if item['word_index'] + 1 < len(offset_mapping):
                    end_idx = offset_mapping[item['word_index'] + 1][0]
                else:
                    end_idx = offset_mapping[item['word_index']][1]
                
                entity_text = text[start_idx:end_idx]
                span_key = (start_idx, end_idx)
            
            # Skip if we've seen this exact span
            if span_key in seen_spans:
                continue
            
            seen_spans.add(span_key)
            
            # Add to entities list
            if len(entity_text.strip()) > 0:
                entities.append({
                    "text": entity_text,
                    "label": item['entity_group'] if 'entity_group' in item else item['entity'],
                    "start": span_key[0],
                    "end": span_key[1],
                    "confidence": item['score']
                })
        
        # Add aviation domain-specific entity extraction
        aviation_entities = self._extract_aviation_entities(text)
        for entity in aviation_entities:
            if (entity["start"], entity["end"]) not in seen_spans:
                entities.append(entity)
                seen_spans.add((entity["start"], entity["end"]))
        
        return entities
    
    def _extract_aviation_entities(self, text: str) -> List[Dict[str, Any]]:
        """
        Extract aviation domain-specific entities using regex patterns.
        
        Args:
            text: Document text
        
        Returns:
            List of extracted aviation entities
        """
        import re
        
        entities = []
        
        # Flight patterns (e.g., ILS RWY 27L)
        ils_pattern = re.compile(r'\b(ILS|VOR|RNAV|GPS|NDB|LOC)\s+RWY\s+(\d{1,2}[LCR]?)\b')
        for match in ils_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "APPROACH_PROCEDURE",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.95
            })
        
        # Altitude patterns (e.g., FL350, 10,000ft)
        alt_pattern = re.compile(r'\b(FL\s*\d{2,3}|\d{1,2},?\d{3}\s*(?:ft|feet))\b')
        for match in alt_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "ALTITUDE",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.95
            })
        
        # Aircraft type designators (e.g., B738, A320)
        aircraft_pattern = re.compile(r'\b([AB][0-9]{3}(?:-[0-9]{1,3})?|B7[0-9]{2}|A3[0-9]{2}|CRJ[0-9]{3}|E[0-9]{3})\b')
        for match in aircraft_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "AIRCRAFT_TYPE",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.92
            })
        
        # FAA regulations (e.g., 14 CFR Part 91, FAR 121.333)
        faa_pattern = re.compile(r'\b(14\s*CFR\s*(?:Part\s*)?|FAR\s*)(\d+)(?:\.(\d+))?\b')
        for match in faa_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "REGULATION_FAA",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.97
            })
        
        # EASA regulations (e.g., CS-25, AMC 25.1309)
        easa_pattern = re.compile(r'\b(CS|AMC|GM)-(\d+)(?:\.(\d+))?\b')
        for match in easa_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "REGULATION_EASA",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.96
            })
        
        # Training procedures (e.g., V1 cut, rejected takeoff)
        training_pattern = re.compile(r'\b(V1\s+cut|RTO|rejected\s+takeoff|go[- ]around|touch[- ]and[- ]go|steep\s+turns|stall\s+recovery)\b', re.IGNORECASE)
        for match in training_pattern.finditer(text):
            entities.append({
                "text": match.group(0),
                "label": "TRAINING_PROCEDURE",
                "start": match.start(),
                "end": match.end(),
                "confidence": 0.90
            })
        
        return entities
    
    def extract_relationships(self, text: str, entities: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Extract relationships between entities in the document.
        
        Args:
            text: Document text
            entities: List of extracted entities
        
        Returns:
            List of relationships between entities
        """
        relationships = []
        
        # Group entities by type
        entity_by_type = {}
        for entity in entities:
            entity_type = entity["label"]
            if entity_type not in entity_by_type:
                entity_by_type[entity_type] = []
            entity_by_type[entity_type].append(entity)
        
        # Extract regulatory relationships (regulations --> procedures)
        if "REGULATION_FAA" in entity_by_type and "TRAINING_PROCEDURE" in entity_by_type:
            for reg in entity_by_type["REGULATION_FAA"]:
                for proc in entity_by_type["TRAINING_PROCEDURE"]:
                    # Check if they are in proximity (within 500 chars)
                    if abs(reg["start"] - proc["start"]) < 500:
                        # Use QA model to check relationship
                        question = f"Is {proc['text']} related to regulation {reg['text']}?"
                        context = text[max(0, min(reg["start"], proc["start"]) - 250):
                                      min(len(text), max(reg["end"], proc["end"]) + 250)]
                        
                        try:
                            answer = self.qa_pipeline(question=question, context=context)
                            if answer["score"] > 0.7:
                                relationships.append({
                                    "source": reg["text"],
                                    "source_type": reg["label"],
                                    "target": proc["text"],
                                    "target_type": proc["label"],
                                    "relationship": "REGULATES",
                                    "confidence": answer["score"]
                                })
                        except Exception as e:
                            logger.warning(f"Error in QA for relationship: {str(e)}")
        
        # Extract aircraft-procedure relationships
        if "AIRCRAFT_TYPE" in entity_by_type and "TRAINING_PROCEDURE" in entity_by_type:
            for aircraft in entity_by_type["AIRCRAFT_TYPE"]:
                for proc in entity_by_type["TRAINING_PROCEDURE"]:
                    # Check if they are in proximity
                    if abs(aircraft["start"] - proc["start"]) < 300:
                        relationships.append({
                            "source": aircraft["text"],
                            "source_type": aircraft["label"],
                            "target": proc["text"],
                            "target_type": proc["label"],
                            "relationship": "HAS_PROCEDURE",
                            "confidence": 0.85
                        })
        
        return relationships
    
    def extract_regulatory_references(self, text: str) -> List[Dict[str, Any]]:
        """
        Extract regulatory references from document text.
        
        Args:
            text: Document text
        
        Returns:
            List of regulatory references
        """
        import re
        
        references = []
        
        # FAA regulations
        faa_pattern = re.compile(r'\b(14\s*CFR\s*(?:Part\s*)?|FAR\s*)(\d+)(?:\.(\d+))?\b')
        for match in faa_pattern.finditer(text):
            part = match.group(2)
            section = match.group(3) if match.group(3) else None
            
            # Get context (text around the reference)
            start_ctx = max(0, match.start() - 100)
            end_ctx = min(len(text), match.end() + 100)
            context = text[start_ctx:end_ctx]
            
            ref = {
                "authority": "FAA",
                "reference": match.group(0),
                "part": part,
                "section": section,
                "context": context.strip(),
                "position": match.start()
            }
            references.append(ref)
        
        # EASA regulations
        easa_pattern = re.compile(r'\b(CS|AMC|GM)-(\d+)(?:\.(\d+))?\b')
        for match in easa_pattern.finditer(text):
            part = match.group(2)
            section = match.group(3) if match.group(3) else None
            
            # Get context
            start_ctx = max(0, match.start() - 100)
            end_ctx = min(len(text), match.end() + 100)
            context = text[start_ctx:end_ctx]
            
            ref = {
                "authority": "EASA",
                "reference": match.group(0),
                "part": part,
                "section": section,
                "context": context.strip(),
                "position": match.start()
            }
            references.append(ref)
        
        # ICAO regulations
        icao_pattern = re.compile(r'\bICAO\s+(?:Annex|Doc)\s+(\d+)(?:\.(\d+))?\b')
        for match in icao_pattern.finditer(text):
            part = match.group(1)
            section = match.group(2) if match.group(2) else None
            
            # Get context
            start_ctx = max(0, match.start() - 100)
            end_ctx = min(len(text), match.end() + 100)
            context = text[start_ctx:end_ctx]
            
            ref = {
                "authority": "ICAO",
                "reference": match.group(0),
                "part": part,
                "section": section,
                "context": context.strip(),
                "position": match.start()
            }
            references.append(ref)
        
        return references
    
    def extract_training_elements(self, text: str) -> Dict[str, Any]:
        """
        Extract training elements like learning objectives, competencies, etc.
        
        Args:
            text: Document text
        
        Returns:
            Dictionary with various training elements
        """
        doc = self.nlp(text[:50000])  # Limit size for performance
        
        # Extract sentences for analysis
        sentences = [sent.text.strip() for sent in doc.sents]
        
        # Learning objectives often start with "By the end of this..." or have verbs like "demonstrate"
        learning_obj_keywords = [
            "demonstrate", "identify", "explain", "describe", "perform", 
            "recognize", "apply", "analyze", "evaluate", "create",
            "objective", "will be able to", "should be able to"
        ]
        
        learning_objectives = []
        for sent in sentences:
            sent_lower = sent.lower()
            if any(keyword in sent_lower for keyword in learning_obj_keywords):
                if len(sent) > 10:  # Avoid short fragments
                    learning_objectives.append(sent)
        
        # Competency areas often mentioned with keywords like "competency", "proficiency", etc.
        competency_keywords = [
            "competency", "competence", "proficiency", "skill", "knowledge",
            "ability", "aptitude", "capability", "performance", "standard"
        ]
        
        competencies = []
        for sent in sentences:
            sent_lower = sent.lower()
            if any(keyword in sent_lower for keyword in competency_keywords):
                if len(sent) > 10:
                    competencies.append(sent)
        
        # Extract assessment criteria
        assessment_keywords = [
            "assessment", "criteria", "evaluation", "grading", "scoring",
            "passing", "satisfactory", "unsatisfactory", "proficient"
        ]
        
        assessment_criteria = []
        for sent in sentences:
            sent_lower = sent.lower()
            if any(keyword in sent_lower for keyword in assessment_keywords):
                if len(sent) > 10:
                    assessment_criteria.append(sent)
        
        return {
            "learning_objectives": learning_objectives[:10],  # Limit to top 10
            "competencies": competencies[:10],
            "assessment_criteria": assessment_criteria[:10]
        }
    
    def extract_document_structure(self, text: str) -> Dict[str, Any]:
        """
        Extract document structure (sections, headings, etc.)
        
        Args:
            text: Document text
        
        Returns:
            Dictionary with document structure information
        """
        # Use heuristics to identify section headings
        import re
        
        lines = text.split('\n')
        structure = {
            "sections": [],
            "toc_detected": False
        }
        
        # Patterns for section headings
        heading_patterns = [
            # Numbered sections: "1. Introduction", "1.2 Requirements", etc.
            r'^(\d+(\.\d+)*)\s+([A-Z][A-Za-z0-9\s\-:]+),
            # Uppercase headings: "INTRODUCTION", "SYSTEM REQUIREMENTS", etc.
            r'^([A-Z][A-Z\s\-:]{3,}),
            # Mixed case headings with colon: "Section: Details"
            r'^([A-Z][A-Za-z0-9\s\-]+):',
            # Title case headings: "Introduction To The System"
            r'^([A-Z][a-z]+(\s+[A-Z][a-z]+)+)
        ]
        
        current_section = None
        section_content = []
        toc_indicators = ["table of contents", "contents", "toc"]
        in_toc = False
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
                
            # Check if this is a table of contents
            if any(indicator in line.lower() for indicator in toc_indicators):
                structure["toc_detected"] = True
                in_toc = True
                continue
                
            # End of TOC detection (a line with page number pattern)
            if in_toc and re.search(r'\d+, line):
                continue
                
            # If we see a line that doesn't match TOC pattern, end TOC section
            if in_toc and not re.search(r'\.{2,}\s*\d+, line):
                in_toc = False
            
            # Check if this is a heading
            is_heading = False
            for pattern in heading_patterns:
                match = re.match(pattern, line)
                if match:
                    # If we've been collecting content for a section, save it
                    if current_section:
                        current_section["content"] = "\n".join(section_content)
                        structure["sections"].append(current_section)
                    
                    # Start new section
                    current_section = {
                        "heading": line,
                        "level": 1 if '.' not in line else line.count('.') + 1,
                        "position": i
                    }
                    section_content = []
                    is_heading = True
                    break
            
            if not is_heading and current_section:
                section_content.append(line)
        
        # Add the last section
        if current_section:
            current_section["content"] = "\n".join(section_content)
            structure["sections"].append(current_section)
        
        return structure
    
    def generate_summary(self, text: str) -> Dict[str, Any]:
        """
        Generate a summary of the document.
        
        Args:
            text: Document text
        
        Returns:
            Dictionary with summary information
        """
        # For simplicity, we'll use a extractive approach
        # In a real implementation, we would use a proper summarization model
        
        doc = self.nlp(text[:50000])  # Limit size for performance
        
        # Extract sentences
        sentences = [sent.text.strip() for sent in doc.sents]
        
        # Filter out short sentences and headers
        filtered_sentences = [s for s in sentences if len(s) > 30 and len(s.split()) > 5]
        
        # If we have very few sentences, return them all
        if len(filtered_sentences) <= 5:
            return {
                "summary": " ".join(filtered_sentences),
                "key_points": filtered_sentences,
                "extracted_sentences": len(filtered_sentences)
            }
        
        # Compute sentence embeddings
        embeddings = self.sentence_model.encode(filtered_sentences)
        
        # Compute centroid of all sentences as document embedding
        doc_embedding = np.mean(embeddings, axis=0)
        
        # Compute similarity of each sentence to document centroid
        similarities = []
        for i, sent_embedding in enumerate(embeddings):
            similarity = np.dot(sent_embedding, doc_embedding) / (
                np.linalg.norm(sent_embedding) * np.linalg.norm(doc_embedding))
            similarities.append((i, similarity))
        
        # Sort by similarity
        similarities.sort(key=lambda x: x[1], reverse=True)
        
        # Take top 5 sentences
        top_indices = [x[0] for x in similarities[:5]]
        top_indices.sort()  # Sort by original order
        
        summary_sentences = [filtered_sentences[i] for i in top_indices]
        
        return {
            "summary": " ".join(summary_sentences),
            "key_points": summary_sentences,
            "extracted_sentences": len(summary_sentences)
        }

    def batch_process_documents(self, file_paths: List[Union[str, Path]], 
                               options: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        Process multiple documents in batch.
        
        Args:
            file_paths: List of paths to document files
            options: Processing options and parameters
        
        Returns:
            List of processed document data
        """
        results = []
        for file_path in file_paths:
            try:
                result = self.process_document(file_path, options)
                results.append(result)
            except Exception as e:
                logger.error(f"Error processing document {file_path}: {str(e)}")
                results.append({
                    "file_path": str(file_path),
                    "error": str(e),
                    "success": False
                })
        
        return results

# Command-line interface
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="AI Document Processor for Advanced Pilot Training Platform")
    parser.add_argument("--input", "-i", required=True, help="Input file or directory to process")
    parser.add_argument("--output", "-o", help="Output file for results (JSON)")
    parser.add_argument("--model-dir", "-m", default="./models", help="Directory containing model files")
    parser.add_argument("--cpu", action="store_true", help="Force CPU usage even if GPU is available")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        processor = AIDocumentProcessor(args.model_dir, not args.cpu)
        
        input_path = Path(args.input)
        if input_path.is_dir():
            # Process all documents in directory
            file_paths = []
            for ext in [".pdf", ".docx", ".doc", ".xlsx", ".xls", ".html", ".htm", ".txt", ".pptx", ".ppt"]:
                file_paths.extend(list(input_path.glob(f"*{ext}")))
            
            results = processor.batch_process_documents(file_paths)
        else:
            # Process single document
            results = [processor.process_document(input_path)]
        
        # Output results
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)
        else:
            print(json.dumps(results, indent=2))
        
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        logger.error(traceback.format_exc())
        sys.exit(1)