import os
import sys
import logging
import json
import time
from typing import List, Dict, Any, Optional, Tuple, Union
from pathlib import Path
import traceback

import numpy as np
import pandas as pd
import torch
from transformers import (
    AutoTokenizer, 
    AutoModelForTokenClassification,
    AutoModelForSequenceClassification,
    AutoModelForQuestionAnswering,
    pipeline
)
from sentence_transformers import SentenceTransformer
import spacy
import pytesseract
from pdf2image import convert_from_path, convert_from_bytes
from pdfminer.high_level import extract_text
import docx2txt
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import fitz  # PyMuPDF

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class AIDocumentProcessor:
    """
    AI-based document processing module for the Advanced Pilot Training Platform.
    
    This module provides capabilities for document understanding, text extraction,
    entity recognition, and relationship extraction from training documents.
    """
    
    def __init__(self, model_dir: str = "./models", use_gpu: bool = True):
        """
        Initialize the AI document processor with the specified models.
        
        Args:
            model_dir: Directory containing pre-trained models
            use_gpu: Whether to use GPU for inference if available
        """
        self.model_dir = Path(model_dir)
        self.use_gpu = use_gpu and torch.cuda.is_available()
        self.device = torch.device("cuda" if self.use_gpu else "cpu")
        
        logger.info(f"Initializing AI Document Processor (using {'GPU' if self.use_gpu else 'CPU'})")
        
        # Load NLP components
        self._load_models()
    
    def _load_models(self):
        """Load all required NLP models."""
        try:
            # Load NER model for aviation/training domain entities
            ner_model_path = self.model_dir / "ner"
            if ner_model_path.exists():
                self.ner_tokenizer = AutoTokenizer.from_pretrained(ner_model_path)
                self.ner_model = AutoModelForTokenClassification.from_pretrained(ner_model_path).to(self.device)
                self.ner_pipeline = pipeline("ner", model=self.ner_model, tokenizer=self.ner_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic NER model
                logger.warning(f"Custom NER model not found at {ner_model_path}, using default model")
                self.ner_pipeline = pipeline("ner", model="dbmdz/bert-large-cased-finetuned-conll03-english", device=0 if self.use_gpu else -1)
            
            # Load document classification model
            classifier_model_path = self.model_dir / "classifier"
            if classifier_model_path.exists():
                self.classifier_tokenizer = AutoTokenizer.from_pretrained(classifier_model_path)
                self.classifier_model = AutoModelForSequenceClassification.from_pretrained(classifier_model_path).to(self.device)
                self.classifier_pipeline = pipeline("text-classification", model=self.classifier_model, tokenizer=self.classifier_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic classification model
                logger.warning(f"Custom classifier model not found at {classifier_model_path}, using default model")
                self.classifier_pipeline = pipeline("text-classification", model="distilbert-base-uncased-finetuned-sst-2-english", device=0 if self.use_gpu else -1)
            
            # Load QA model for relation extraction
            qa_model_path = self.model_dir / "qa"
            if qa_model_path.exists():
                self.qa_tokenizer = AutoTokenizer.from_pretrained(qa_model_path)
                self.qa_model = AutoModelForQuestionAnswering.from_pretrained(qa_model_path).to(self.device)
                self.qa_pipeline = pipeline("question-answering", model=self.qa_model, tokenizer=self.qa_tokenizer, device=0 if self.use_gpu else -1)
            else:
                # Fallback to generic QA model
                logger.warning(f"Custom QA model not found at {qa_model_path}, using default model")
                self.qa_pipeline = pipeline("question-answering", model="deepset/roberta-base-squad2", device=0 if self.use_gpu else -1)
            
            # Load sentence embedding model for similarity and clustering
            self.sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
            if self.use_gpu:
                self.sentence_model = self.sentence_model.to(self.device)
            
            # Load spaCy for text processing
            try:
                self.nlp = spacy.load("en_core_web_trf")
            except OSError:
                logger.warning("Could not load en_core_web_trf. Downloading now...")
                spacy.cli.download("en_core_web_trf")
                self.nlp = spacy.load("en_core_web_trf")
            
            logger.info("All NLP models loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading models: {str(e)}")
            logger.error(traceback.format_exc())
            raise RuntimeError(f"Failed to initialize AI Document Processor: {str(e)}")
    
    def process_document(self, file_path: Union[str, Path], 
                        options: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Process a document file and extract structured information.
        
        Args:
            file_path: Path to the document file
            options: Processing options and parameters
        
        Returns:
            Dictionary containing processed document data
        """
        if options is None:
            options = {}
        
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Document file not found: {file_path}")
        
        try:
            # Determine document type and extract text
            doc_type, text = self.extract_text(file_path)
            
            # Process the text
            doc_info = {
                "document_type": doc_type,
                "text": text[:1000] + "..." if len(text) > 1000 else text,  # Truncate for result object
                "file_path": str(file_path),
                "file_name": file_path.name,
                "file_size": file_path.stat().st_size,
                "processing_time": time.time(),
            }
            
            # Document classification
            if options.get("classify_document", True):
                doc_info["document_classification"] = self.classify_document(text)
            
            # Extract entities
            if options.get("extract_entities", True):
                doc_info["entities"] = self.extract_entities(text)
            
            # Extract relationships
            if options.get("extract_relationships", True):
                doc_info["relationships"] = self.extract_relationships(text, doc_info.get("entities", []))
            
            # Extract regulatory references
            if options.get("extract_regulations", True):
                doc_info["regulatory_references"] = self.extract_regulatory_references(text)
            
            # Extract training elements
            if options.get("extract_training_elements", True):
                doc_info["training_elements"] = self.extract_training_elements(text)
            
            # Extract document structure
            if options.get("extract_structure", True):
                doc_info["structure"] = self.extract_document_structure(text)
            
            # Generate summary
            if options.get("generate_summary", True):
                doc_info["summary"] = self.generate_summary(text)
            
            return doc_info
            
        except Exception as e:
            logger.error(f"Error processing document {file_path}: {str(e)}")
            logger.error(traceback.format_exc())
            raise RuntimeError(f"Failed to process document: {str(e)}")
    
    def extract_text(self, file_path: Path) -> Tuple[str, str]:
        """
        Extract text from a document file based on its type.
        
        Args:
            file_path: Path to the document file
        
        Returns:
            Tuple of (document_type, extracted_text)
        """
        suffix = file_path.suffix.lower()
        
        if suffix == '.pdf':
            return "pdf", self._extract_from_pdf(file_path)
        elif suffix in ['.docx', '.doc']:
            return "docx", self._extract_from_docx(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return "excel", self._extract_from_excel(file_path)
        elif suffix in ['.html', '.htm']:
            return "html", self._extract_from_html(file_path)
        elif suffix == '.txt':
            return "text", file_path.read_text(errors='replace')
        elif suffix in ['.pptx', '.ppt']:
            return "powerpoint", self._extract_from_ppt(file_path)
        else:
            # Try to extract as text as a fallback
            try:
                return "unknown", file_path.read_text(errors='replace')
            except Exception:
                raise ValueError(f"Unsupported document type: {suffix}")
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from a PDF file using multiple methods for best results."""
        text = ""
        
        # Try pdfminer first (better for text-based PDFs)
        try:
            text = extract_text(str(file_path))
        except Exception as e:
            logger.warning(f"pdfminer extraction failed: {str(e)}")
        
        # If text is too short, try PyMuPDF
        if len(text.strip()) < 100:
            try:
                doc = fitz.open(str(file_path))
                pages_text = []
                for page in doc:
                    pages_text.append(page.get_text())
                text = "\n\n".join(pages_text)
                doc.close()
            except Exception as e:
                logger.warning(f"PyMuPDF extraction failed: {str(e)}")
        
        # If still too short, try OCR
        if len(text.strip()) < 100:
            try:
                images = convert_from_path(str(file_path))
                ocr_texts = []
                for img in images:
                    ocr_texts.append(pytesseract.image_to_string(img))
                text = "\n\n".join(ocr_texts)
            except Exception as e:
                logger.warning(f"OCR extraction failed: {str(e)}")
        
        return text
    
    def _extract_from_docx(self, file_path: Path) -> str:
        """Extract text from a Word document."""
        return docx2txt.process(str(file_path))
    
    def _extract_from_excel(self, file_path: Path) -> str:
        """Extract text from an Excel spreadsheet."""
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        texts = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            texts.append(f"Sheet: {sheet}")
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            
            texts.append("\n".join(rows))
        
        wb.close()
        return "\n\n".join(texts)
    
    def _extract_from_html(self, file_path: Path) -> str:
        """Extract text from an HTML file."""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            return soup.get_text(separator="\n")
    
    def _extract_from_ppt(self, file_path: Path) -> str:
        """Extract text from a PowerPoint presentation."""
        # This is a placeholder - in a real implementation we would use
        # python-pptx or a similar library
        return f"PowerPoint extraction not fully implemented for {file_path}"
    
    def classify_document(self, text: str) -> Dict[str, Any]:
        """
        Classify the document type and category.
        
        Args:
            text: Document text
        
        Returns:
            Dictionary with classification results
        """
        # Use only first 1000 tokens for classification
        truncated_text = " ".join(text.split()[:1000])
        
        # Get predicted categories
        result = self.classifier_pipeline(truncated_text)
        
        # Convert result to our expected format
        if isinstance(result, list):
            result = result[0]
        
        return {
            "category": result["label"],
            "confidence": float(result["score"])
        }
    
    def extract_entities(self, text: str) -> List[Dict[str, Any]]:
        """
        Extract named entities from document text.
        
        Args:
            text: Document text
        
        Returns:
            List of extracted entities with metadata
        """
        # Use spaCy for initial entity extraction
        doc = self.nlp(text[:10000])  # Limit size for performance
        
        entities = []
        for ent in doc.ents:
            entities.append({
                "text": ent.text,
                "label": ent.label_,
                "start": ent.start_char,
                "end": ent.end_char,
                "confidence": 0.9  # Placeholder, spaCy doesn't provide confidence scores
            })
        
        # Extract domain-specific entities using the specialized NER model
        # Process text in chunks of 512 tokens to avoid exceeding model length limits
        tokenized = self.ner_tokenizer(text, return_offsets_mapping=True, truncation=True)
        offset_mapping = tokenized["offset_mapping"]
        
        specialized_results = self.ner_pipeline(text[:10000])
        
        # Process and deduplicate NER results
        seen_spans = set()
        for item in specialized_results:
            # Map subtoken indices to character spans
            if 'start' in item and 'end' in item:
                entity_text = text[item['start']:item['end']]
                span_key = (item['start'], item['end'])
            else: